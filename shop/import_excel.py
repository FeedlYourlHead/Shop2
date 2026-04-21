import openpyxl
from django.db import IntegrityError
from .models import Category, Product
from pytils.translit import slugify

def import_from_excel(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    categories_created = 0
    products_created = 0
    errors = []

    if 'category' in wb.sheetnames:
        ws = wb['category']
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
